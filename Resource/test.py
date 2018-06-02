#to read from file
from openpyxl import *
sb = load_workbook('test.xlsm', read_only = False,  keep_vba = True)
ss = sb['Sheet1']
count = 2
#print(ss['A' + str(count)].value)
while(ss['A' + str(count)].value != None):
	ss['A' + str(count)] = None
	count += 1
	if (count == 100):
		print("Ended potentinally endless loop")
sb.save('test2.xlsm')
"""
wb = Workbook()
ss = sb['Sheet1']
ws = wb.active
ws['B1'] = "Start Time"
ws['C1'] = "Duration"
count = 2
while(ws['A' + str(count)].value != None):
	print(ws['A' + str(count)].value)
	count += 1
wb.save('test2.xlsm')
"""