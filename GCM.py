print("Running application")
import functools
print("Importing openpyxl")
from openpyxl import Workbook, load_workbook
print("Importing subprocess")
import subprocess
print("Importing sys")
print("Importing PyQt5")
from PyQt5.QtWidgets import *
from PyQt5.QtCore import QDate, QTime
class FileIO():
	def openfile(filename):
		try:
			proc = subprocess.Popen(filename, shell=True, stdout=subprocess.PIPE)
			proc.wait()
			print("The programme exits with return code of:")
			print(proc.returncode)
		except:
			print("An error occured.")
	def createBlank():
		sb = load_workbook('Resource/file.xlsm', read_only = False, keep_vba = True)
		ss = sb['Sheet1']
		count = 2
		#print(ss['A' + str(count)].value)
		while(ss['A' + str(count)].value != None):
			ss['A' + str(count)] = None
			ss['B' + str(count)] = None
			ss['C' + str(count)] = None
			count += 1
			if (count == 100):
				print("Ended potentionally endless loop")
		sb.save('output.xlsm')
	def insertValues(values):
		sb = load_workbook('output.xlsm', read_only = False, keep_vba = True)
		ss = sb['Sheet1']
		for i in range(len(values)):
#			print(values[i].returnDateObj().toString('dd/MM/yyyy'))
			ss['A' + str(i + 2)] = values[i].returnTaskName()
			ss['B' + str(i + 2)] = values[i].returnDateObj().toString('dd/MM/yyyy')
			ss['C' + str(i + 2)] = values[i].returnDayDuration()
		sb.save('output.xlsm')
class Tasks:
	taskCount = 0
	def __init__(self, dat, ddu, tms, tme, tan, pri):
		Tasks.taskCount += 1
		self.date = dat
		self.dayDuration = ddu
		self.timeStart = tms
		self.timeEnd = tme
		self.priority = pri
		if (dat == None):
			self.date = QDate(2001, 1, 1)
		if (ddu == None):
			self.dayDuration = 0
		if (tms == None):
			self.timeStart = QTime(0,0,0,0)
		if (tme == None):
			self.timeEnd = QTime(23,59,0,0)
		if (pri == None):
			self.priority = 0
		if (tan == ""):
			self.taskName = "Task " + str(Tasks.taskCount)
		else:
			self.taskName = tan
	def toString(self):
		toReturn = self.taskName + ": " + self.date.toString() + " - " + self.date.addDays(self.dayDuration).toString()
		if (self.priority == 0):
			toReturn += ", Low Priority"
		elif (self.priority == 1):
			toReturn += ", Medium Priority"
		else:
			toReturn += ", High Priority"
		return toReturn
	def returnDateObj(self):
		return self.date
	def returnTaskName(self):
		return self.taskName
	def returnDayDuration(self):
		return self.dayDuration
class Applic(QMainWindow):
	def __init__(self):
		super().__init__()
		self.tasksList = []
		self.eventName = "Untitled"
		self.taskName = ""
		self.__initUI__()

	def __initUI__(self):
		self.setMenuBar()
		QLabel("Task Name: ", self).move(25,25)
		self.taskNameInput = QLineEdit(self)
		self.taskNameInput.move(90,27)
		self.taskNameInput.resize(450,25)
		self.taskNameInput.textChanged.connect(self.varTaskName)
		self.createTaskButton = self.createButton("Create New Task", 550,28, functools.partial(self.createTaskBtnClk))
		self.statusBar().showMessage('Ready')

		#main window configuration
		self.resize(700,420)
		qr = self.frameGeometry()
		cp = QDesktopWidget().availableGeometry().center()
		qr.moveCenter(cp)
		self.move(qr.topLeft())
		self.setWindowTitle("YSimple - " + self.eventName + ".xlsm")
		self.show()

	def setMenuBar(self):
		createNewEventAct = QAction("&New", self)
		createNewEventAct.setShortcut("Ctrl+N")
		createNewEventAct.setStatusTip("Create new event")
		createNewEventAct.triggered.connect(self.createNewEventBtnClk)
		
		openEvent = QAction("&Open", self)
		openEvent.setShortcut("Ctrl+O")
		openEvent.setStatusTip("Open existing event")
		openEvent.triggered.connect(self.empty)

		saveEvent = QAction("&Save", self)
		saveEvent.setShortcut("Ctrl+S")
		saveEvent.setStatusTip("Save current event")
		saveEvent.triggered.connect(self.commitWork)
		
		saveNewEvent = QAction("&Save as", self)
		saveNewEvent.setShortcut("Ctrl+Shift+S")
		saveNewEvent.setStatusTip("Save current event as new file")
		saveNewEvent.triggered.connect(self.empty)

		exitApp = QAction("&Save as", self)
		exitApp.setShortcut("Ctrl+W")
		exitApp.setStatusTip("Save current event as new file")
		exitApp.triggered.connect(qApp.quit)
		
		menubar = self.menuBar()
		fileMenu = menubar.addMenu("&File")
		fileMenu.addAction(createNewEventAct)
		fileMenu.addAction(openEvent)
		fileMenu.addAction(saveEvent)
		fileMenu.addAction(saveNewEvent)
		fileMenu.addAction(exitApp)

	def createNewEventBtnClk(self):
		msgBox = QMessageBox()
		msgBox.setIcon(QMessageBox.Question)
		msgBox.setText("Current event is not empty.")
		msgBox.setInformativeText("Continue?")
		msgBox.setStandardButtons(QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel)
		ret = msgBox.exec_()
		if (ret == QMessageBox.Save):
			print("Project saved.")
		elif (ret == QMessageBox.Discard):
			print("New project is created.")
		else:
			pass

	def createTaskBtnClk(self):
		self.createTask(None, None, None, None, self.taskName, None)
		self.allEntryClear()

	def allEntryClear(self):
		self.taskNameInput.clear()

	def varTaskName(self, text):
		self.taskName = text
		
	def createTask(self, dat, ddu, tms, tme, tan, pri):
		# if (dat == None or ddu == None or tms == None or tme == None or tan == None or pri == None):
			# print("Default value for the parameters is used.")
			# self.tasksList.append(Tasks())
		# else:
		self.tasksList.append(Tasks(dat, ddu, tms, tme, tan, pri))
		self.statusBar().showMessage("Task created: " + self.tasksList[len(self.tasksList) - 1].toString())
	
	def deleteTask(self, taskNumber = None):
		if (taskNumber == None):
			print("Deleting the last task.")
			taskNumber = len(self.tasksList)
		else:
			print("Deleting task number " + str(taskNumber))
		try:
			self.tasksList.remove(self.tasksList[taskNumber-1])
		except:
			print("Deletion not successful")
	def commitWork(self):
		FileIO.createBlank()
		FileIO.insertValues(self.tasksList)
	
	def createButton(self, text, posx, posy, taskFunction):
		btn = QPushButton(text, self)
		btn.setToolTip(text)
		btn.resize(btn.sizeHint())
		btn.move(posx, posy)
		btn.clicked.connect(taskFunction)
		return btn
	def empty(self):
		print('Action triggered.')
try:
	if __name__ == '__main__':
		app = QApplication([])
		ex = Applic()
		app.exec_()
except:
	print('Error')
#FileIO.openfile("output.xlsm")